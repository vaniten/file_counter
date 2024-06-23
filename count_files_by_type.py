import os
import csv
import sys
from datetime import datetime, timedelta
from PyPDF2 import PdfReader
import docx
import xlrd
from pptx import Presentation
from openpyxl import load_workbook
from prettytable import PrettyTable

def count_pages_pdf(file_path):
    try:
        reader = PdfReader(file_path)
        return len(reader.pages)
    except Exception as e:
        print(f"Error reading PDF {file_path}: {e}")
        return 0

def count_pages_doc(file_path):
    try:
        with open(file_path, 'rb') as f:
            data = f.read()
            return data.count(b'\f') + 1
    except Exception as e:
        print(f"Error reading DOC {file_path}: {e}")
        return 0

def count_pages_docx(file_path):
    try:
        doc = docx.Document(file_path)
        return len(doc.element.xpath('//w:sectPr'))
    except Exception as e:
        print(f"Error reading DOCX {file_path}: {e}")
        return 0

def count_pages_pptx(file_path):
    try:
        prs = Presentation(file_path)
        return len(prs.slides)
    except Exception as e:
        print(f"Error reading PPTX {file_path}: {e}")
        return 0

def count_rows_columns_xls(file_path):
    try:
        workbook = xlrd.open_workbook(file_path)
        total_rows = 0
        total_columns = 0
        for sheet in workbook.sheets():
            total_rows += sheet.nrows
            total_columns += sheet.ncols
        return total_rows, total_columns
    except Exception as e:
        print(f"Error reading XLS {file_path}: {e}")
        return 0, 0

def count_rows_columns_xlsx(file_path):
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        total_rows = 0
        total_columns = 0
        for sheet in workbook:
            max_row = sheet.max_row or 0
            max_column = sheet.max_column or 0
            total_rows += max_row
            total_columns += max_column
        return total_rows, total_columns
    except Exception as e:
        print(f"Error reading XLSX {file_path}: {e}")
        return 0, 0

def update_page_ranges(page_ranges, pages):
    if pages <= 2:
        page_ranges['1-2'] += 1
    elif pages <= 5:
        page_ranges['3-5'] += 1
    else:
        page_ranges['>5'] += 1

def main(directory, output_file="output.csv"):
    start_time = datetime.now()

    file_counts = {}
    file_sizes = {}
    total_pages = {}
    total_rows = {}
    total_columns = {}
    page_ranges = {
        'pdf': {'1-2': 0, '3-5': 0, '>5': 0},
        'doc': {'1-2': 0, '3-5': 0, '>5': 0},
        'docx': {'1-2': 0, '3-5': 0, '>5': 0},
        'pptx': {'1-2': 0, '3-5': 0, '>5': 0}
    }
    page_ranges_by_year = {}

    files_to_process = []
    for root, _, files in os.walk(directory):
        for file in files:
            files_to_process.append((root, file))

    total_files = len(files_to_process)
    for i, (root, file) in enumerate(files_to_process):
        print(f"Processing file {i + 1} of {total_files}", end='\r')

        ext = file.split('.')[-1].lower()
        file_path = os.path.join(root, file)
        file_size = os.path.getsize(file_path)

        try:
            modified_time = os.path.getmtime(file_path)
            modified_date = datetime.fromtimestamp(modified_time)
            year = modified_date.year
        except Exception as e:
            print(f"Error getting modified date for {file_path}: {e}")
            year = None

        if year is not None:
            file_counts.setdefault((ext, year), 0)
            file_sizes.setdefault((ext, year), 0)
            total_pages.setdefault((ext, year), 0)
            total_rows.setdefault((ext, year), 0)
            total_columns.setdefault((ext, year), 0)
            if (ext, year) not in page_ranges_by_year:
                page_ranges_by_year[(ext, year)] = {'1-2': 0, '3-5': 0, '>5': 0}

            file_counts[(ext, year)] += 1
            file_sizes[(ext, year)] += file_size

            pages = 0
            if ext == "pdf":
                pages = count_pages_pdf(file_path)
            elif ext == "doc":
                pages = count_pages_doc(file_path)
            elif ext == "docx":
                pages = count_pages_docx(file_path)
            elif ext == "pptx":
                pages = count_pages_pptx(file_path)

            if pages > 0:
                total_pages[(ext, year)] += pages
                update_page_ranges(page_ranges_by_year[(ext, year)], pages)
                if ext in page_ranges:
                    update_page_ranges(page_ranges[ext], pages)

            if ext == "xls":
                rows, columns = count_rows_columns_xls(file_path)
                total_rows[(ext, year)] += rows
                total_columns[(ext, year)] += columns
            elif ext == "xlsx":
                rows, columns = count_rows_columns_xlsx(file_path)
                total_rows[(ext, year)] += rows
                total_columns[(ext, year)] += columns

    rows = []
    for (ext, year), count in file_counts.items():
        page_range = page_ranges_by_year.get((ext, year), {'1-2': 0, '3-5': 0, '>5': 0})
        row = [
            ext, year, count, file_sizes[(ext, year)], total_pages.get((ext, year), 0),
            total_rows.get((ext, year), 0), total_columns.get((ext, year), 0),
            page_range['1-2'], page_range['3-5'], page_range['>5']
        ]
        rows.append(row)

    # Group rows by year and sort by count within each year (descending)
    rows_by_year = {}
    for row in rows:
        year = row[1]
        if year not in rows_by_year:
            rows_by_year[year] = []
        rows_by_year[year].append(row)

    for year in rows_by_year:
        rows_by_year[year].sort(key=lambda x: -x[2])  # Sort by count (descending)

    summary_totals = {}
    total_summary_count = 0
    total_summary_size = 0
    total_summary_pages = 0
    total_summary_rows = 0
    total_summary_columns = 0
    total_summary_1_2_pages = 0
    total_summary_3_5_pages = 0
    total_summary_more_than_5_pages = 0

    for year in rows_by_year:
        for row in rows_by_year[year]:
            ext = row[0]
            summary_totals.setdefault(ext, {"count": 0, "size": 0, "pages": 0, "rows": 0, "columns": 0, "1-2": 0, "3-5": 0, ">5": 0})
            summary_totals[ext]["count"] += row[2]
            summary_totals[ext]["size"] += row[3]
            summary_totals[ext]["pages"] += row[4]
            summary_totals[ext]["rows"] += row[5]
            summary_totals[ext]["columns"] += row[6]
            summary_totals[ext]["1-2"] += row[7]
            summary_totals[ext]["3-5"] += row[8]
            summary_totals[ext][">5"] += row[9]
            total_summary_count += row[2]
            total_summary_size += row[3]
            total_summary_pages += row[4]
            total_summary_rows += row[5]
            total_summary_columns += row[6]
            total_summary_1_2_pages += row[7]
            total_summary_3_5_pages += row[8]
            total_summary_more_than_5_pages += row[9]

    # Create summary table sorted by count descending
    sorted_summary_totals = sorted(summary_totals.items(), key=lambda item: -item[1]['count'])

    summary_table = PrettyTable()
    summary_table.field_names = ["File Type", "Count", "Total Size (bytes)", "Total Pages", "Total Rows", "Total Columns", "1-2 Pages", "3-5 Pages", ">5 Pages"]

    with open(output_file, mode='w', newline='') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["Summary Totals"])
        writer.writerow(["File Type", "Count", "Total Size (bytes)", "Total Pages", "Total Rows", "Total Columns", "1-2 Pages", "3-5 Pages", ">5 Pages"])

        for ext, totals in sorted_summary_totals:
            row = [
                ext, totals["count"], totals["size"], totals["pages"],
                totals["rows"], totals["columns"], totals["1-2"], totals["3-5"], totals[">5"]
            ]
            summary_table.add_row(row)
            writer.writerow(row)

        summary_table.add_row([
            "TOTALS", total_summary_count, total_summary_size, total_summary_pages,
            total_summary_rows, total_summary_columns, total_summary_1_2_pages,
            total_summary_3_5_pages, total_summary_more_than_5_pages
        ])
        writer.writerow([
            "TOTALS", total_summary_count, total_summary_size, total_summary_pages,
            total_summary_rows, total_summary_columns, total_summary_1_2_pages,
            total_summary_3_5_pages, total_summary_more_than_5_pages
        ])

        print("\n\nSummary Table:")
        print(summary_table)

        # Write page range summary table
        page_range_table = PrettyTable()
        page_range_table.field_names = ["File Type", "1-2 Pages", "3-5 Pages", ">5 Pages"]

        writer.writerow([])
        writer.writerow(["Page Range Summary"])
        writer.writerow(["File Type", "1-2 Pages", "3-5 Pages", ">5 Pages"])

        for ext, ranges in page_ranges.items():
            row = [ext, ranges["1-2"], ranges["3-5"], ranges[">5"]]
            page_range_table.add_row(row)
            writer.writerow(row)

        print("\nPage Range Table:")
        print(page_range_table)

        # Write detailed year tables
        writer.writerow([])
        for year in sorted(rows_by_year.keys(), reverse=True):
            year_table = PrettyTable()
            year_table.field_names = ["File Type", "Year", "Count", "Total Size (bytes)", "Total Pages", "Total Rows", "Total Columns", "1-2 Pages", "3-5 Pages", ">5 Pages"]

            year_count = 0
            year_size = 0
            year_pages = 0
            year_rows = 0
            year_columns = 0
            year_1_2_pages = 0
            year_3_5_pages = 0
            year_more_than_5_pages = 0

            for row in rows_by_year[year]:
                year_table.add_row(row)
                writer.writerow(row)
                year_count += row[2]
                year_size += row[3]
                year_pages += row[4]
                year_rows += row[5]
                year_columns += row[6]
                year_1_2_pages += row[7]
                year_3_5_pages += row[8]
                year_more_than_5_pages += row[9]

            year_table.add_row([
                "TOTALS", year, year_count, year_size, year_pages, year_rows, year_columns,
                year_1_2_pages, year_3_5_pages, year_more_than_5_pages
            ])
            writer.writerow([
                "TOTALS", year, year_count, year_size, year_pages, year_rows, year_columns,
                year_1_2_pages, year_3_5_pages, year_more_than_5_pages
            ])

            print(f"\nYear {year} Detail:")
            print(year_table)


    end_time = datetime.now()
    duration = end_time - start_time
    print(f"\nTotal runtime: {duration}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python script.py <directory_to_search> [output_file]")
        sys.exit(1)

    directory_to_search = sys.argv[1]
    output_file_path = sys.argv[2] if len(sys.argv) > 2 else "output.csv"
    main(directory_to_search, output_file_path)
